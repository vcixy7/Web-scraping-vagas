# exportação das vagas para uma planilha Excel usando pandas
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

COLUNAS = ["Título", "Empresa", "Local", "Estado (UF)", "Modalidade", "Salário", "Tecnologias", "Fonte", "Link"]

# largura máxima de uma coluna; o que passar disso quebra em várias linhas
LARGURA_MAXIMA = 50


def salvar_excel(vagas, nome_arquivo="vagas.xlsx"):
    linhas = []
    for vaga in vagas:
        linhas.append({
            "Título": vaga.get("titulo"),
            "Empresa": vaga.get("empresa"),
            "Local": vaga.get("local"),
            "Estado (UF)": vaga.get("estado"),
            "Modalidade": vaga.get("modalidade"),
            "Salário": vaga.get("salario_texto"),
            "Tecnologias": ", ".join(vaga.get("tecnologias", [])),
            "Fonte": vaga.get("fonte"),
            "Link": vaga.get("url"),
        })
    df = pd.DataFrame(linhas, columns=COLUNAS)

    with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vagas")
        planilha = writer.sheets["Vagas"]

        # largura de cada coluna: cabe o texto, com um teto (o resto quebra em linhas)
        for i, coluna in enumerate(df.columns, start=1):
            tamanhos = [len(str(coluna))]
            tamanhos += [len(str(valor)) for valor in df[coluna] if valor is not None]
            largura = min(max(tamanhos) + 2, LARGURA_MAXIMA)
            planilha.column_dimensions[get_column_letter(i)].width = largura

        # quebra de texto em todas as células, para nada ficar cortado
        quebra = Alignment(wrap_text=True, vertical="top")
        for linha in planilha.iter_rows():
            for celula in linha:
                celula.alignment = quebra

        # cabeçalho em negrito
        for celula in planilha[1]:
            celula.font = Font(bold=True)

        # mantém o cabeçalho visível ao rolar a planilha
        planilha.freeze_panes = "A2"
